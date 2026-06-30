"""Generér regel-sporbarhedsmatrixen (xlsx) fra regelkataloget.

Læser customs/rules/Customs-Validation-Rules.json og skriver
docs/Customs-Analytics_Regel-sporbarhedsmatrix.xlsx med fanerne:
  1. Oversigt              — nøgletal (katalogversion, antal regler/lag).
  2. Valideringsregler     — regel → autoritativ kilde → modul → test (kernen).
  3. Referencedata         — provenance for TARIC/eVita + Toldstyrelsens dms-public.

Kør: python tools/build_traceability.py
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CATALOG = ROOT / "customs" / "rules" / "Customs-Validation-Rules.json"
SOURCE_COMMIT = ROOT / "reference" / "SOURCE_COMMIT.txt"
OUT = ROOT / "docs" / "Customs-Analytics_Regel-sporbarhedsmatrix.xlsx"

NAVY = "1B365D"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
CELL_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


def _sheet(wb, title, headers, rows, widths):
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    for r, rowvals in enumerate(rows, start=2):
        for c, v in enumerate(rowvals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = CELL_FONT
            cell.alignment = WRAP
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    return ws


def build():
    catalog = json.loads(CATALOG.read_text(encoding="utf-8"))
    rules = catalog["rules"]
    layers = {l["id"]: l["label_da"] for l in catalog["layers"]}
    sev = {s["id"]: s["label_da"] for s in catalog["severity_levels"]}
    status = {s["id"]: s["label_da"] for s in catalog["statuses"]}
    impl = [r for r in rules if r["status"] == "implemented"]
    plan = [r for r in rules if r["status"] == "planned"]

    src_commit = SOURCE_COMMIT.read_text(encoding="utf-8").strip() if SOURCE_COMMIT.exists() else "(ukendt)"

    wb = Workbook()
    wb.remove(wb.active)

    # 1. Oversigt
    _sheet(wb, "Oversigt", ["Nøgletal", "Værdi"], [
        ["Værktøj", "Customs Analytics"],
        ["Regelkatalog-version", catalog["catalog_version"]],
        ["Katalog frigivet", catalog["catalog_released"]],
        ["Regler i alt", len(rules)],
        ["— implementeret", len(impl)],
        ["— planlagt (udkast)", len(plan)],
        ["Lag", len(catalog["layers"])],
        ["Klassifikation", "Fortroligt — internt"],
    ], [28, 60])

    # 2. Valideringsregler — kernen: regel → kilde → modul → test
    rows = []
    for r in sorted(rules, key=lambda x: (x["layer"], x["id"])):
        rows.append([
            r["id"],
            f"{r['layer']} · {layers[r['layer']]}",
            r["name_da"],
            sev.get(r["severity"], r["severity"]),
            status.get(r["status"], r["status"]),
            r["authoritative_source"],
            r["implementation"],
            r["test"],
        ])
    _sheet(wb, "Valideringsregler",
           ["Regel", "Lag", "Kontrol", "Severity", "Status", "Autoritativ kilde", "Modul", "Test"],
           rows, [10, 28, 34, 20, 18, 42, 40, 36])

    # 3. Referencedata — provenance
    _sheet(wb, "Referencedata", ["Kilde", "Indhold", "Mekanisme / provenance"], [
        ["Toldstyrelsen — skat/dms-public",
         "H1/I1-XSD'er, kodelister, forretningsregler, DMS-datamodel",
         f"Vendret i reference/; kilde-commit: {src_commit}"],
        ["TARIC — eVita-toldtarif (Skattestyrelsen)",
         "MFN-satser + officielle danske varebeskrivelser (15.767 deklarérbare koder)",
         "info.skat.dk/download/told/toldtarif_ext.zip → reference/tariff/mfn_rates.csv (tools/sync_taric.py --evita)"],
        ["TARIC — Trader Export Total",
         "Præferencesatser pr. HS×område + geografiske grupper (temporal)",
         "info.skat.dk/download/told/tot/ → reference/tariff/preferential_rates.csv + geo_areas.json (tools/sync_preferences.py)"],
        ["EU — Access2Markets",
         "FTA-/GSP-/EPA-/toldunions-dækning (krydstjek af præference-grundlag)",
         "Manuelt krydstjekket 2026-06-17; toldunioner (measureType 106) medtaget"],
        ["Vedligehold",
         "Månedlig automatisk opdatering af MFN + præferencer",
         ".github/workflows/update-tariff.yml (den 2. i hver måned) + tools/update_tariff.sh med pytest-selvtjek"],
    ], [38, 46, 70])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Skrev {OUT}  ({len(impl)} implementerede + {len(plan)} planlagte regler)")


if __name__ == "__main__":
    build()
