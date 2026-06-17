"""TARIC-sync — omsætter et officielt TARIC-ekstrakt til tariferingslagets format.

Læser TARIC-databaseekstraktets tre logiske dele og skriver
`reference/tariff/mfn_rates.csv` + `reference/tariff/arrangements.json`, som
`customs/tariff.py` så foretrækker frem for seed'et.

Input (CSV eller XLSX — TARIC-ekstraktet leveres som XLSX fra DG TAXUD/CIRCABC):
- **nomenclature**: varekode (10-cif.), IS_LEAF (1 = deklarérbar), beskrivelse.
- **measures**: varekode, measure_type, geografisk område, toldsats (duty expression).
  measure_type 103 = tredjelandstold (MFN, område ERGA OMNES); 142/145 = præference/GSP.
- **geo_groups**: gruppe → medlemslande + aftalens navn/type (fta/customs_union/
  gsp_plus/eba/gsp_standard). I TARIC identificeres aftalen via område + retsgrundlag;
  her samles det i én gruppefil.

Kør:
    python tools/sync_taric.py --nomenclature N.xlsx --measures M.xlsx --geo-groups G.csv
    python tools/sync_taric.py --check        # verificér at referencedata findes

Begrænsning: kun ad valorem-satser udtrækkes; specifik told (EUR/100 kg) og
sammensatte satser markeres som ikke-modellerede (rate=None) indtil videre.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

REFERENCE = Path(__file__).resolve().parent.parent / "reference" / "tariff"

MFN_MEASURE_TYPES = {"103"}            # tredjelandstold (erga omnes)
PREFERENCE_MEASURE_TYPES = {"142", "145"}  # tariff preference / GSP
ERGA_OMNES = {"1011", "ERGA OMNES", "ERGAOMNES", "ALL"}
PREFERRED_LANGS = ["DA", "EN"]         # beskrivelsessprog (DDS-ekstrakt er flersproget)

# Det officielle DDS TARIC-ekstrakts download-mekanik (verificeret, ingen login).
DDS_BASE = "https://ec.europa.eu/taxation_customs/dds2/taric"

# Kolonne-aliaser (normaliseret) → kanonisk felt. Dækker både det officielle DDS-
# ekstrakt (GOODS CODE, MEAS_TYP_ID, GEOGR_AREA, DUTY, DESCR_TEXT, LANG_COD) og
# det forenklede fixtur-/CIRCABC-format.
_COLS = {
    "goods_code": ["goodscode", "goods_code", "varekode", "code", "goodsnomenclatureitemid"],
    "is_leaf": ["isleaf", "is_leaf", "leaf", "declarable"],
    "description": ["descrtext", "description", "beskrivelse", "desc"],
    "lang": ["langcod", "lang", "language", "languagecode"],
    "measure_type": ["meastypid", "measuretype", "measure_type", "measuretypeid", "type"],
    "geo_area": ["geograrea", "geographicalarea", "geo_area", "geoarea", "area", "geographicalareaid"],
    "duty_rate": ["duty", "dutyrate", "duty_rate", "dutyexpression", "rate", "dutyamount"],
    "group_id": ["groupid", "group_id", "group", "area", "geographicalareaid"],
    "member_country": ["membercountry", "member_country", "member", "country", "countrycode"],
    "arrangement_name": ["arrangementname", "arrangement_name", "name", "agreement"],
    "arrangement_type": ["arrangementtype", "arrangement_type", "type"],
}


def _norm(name: str) -> str:
    return "".join(ch for ch in str(name).lower() if ch.isalnum())


def _read_table(path: Path) -> list[dict]:
    """Læs CSV eller XLSX til en liste af dicts (med originale kolonnenavne)."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(path.read_bytes()), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(it)]
        rows = [
            {header[i]: v[i] for i in range(min(len(header), len(v)))}
            for v in it if v is not None and any(c is not None for c in v)
        ]
        wb.close()
        return rows
    text = path.read_text(encoding="utf-8-sig")
    delim = ";" if text[:2048].count(";") > text[:2048].count(",") else ","
    return list(csv.DictReader(io.StringIO(text), delimiter=delim))


def _colmap(header: list[str]) -> dict[str, str]:
    norm = {_norm(h): h for h in header}
    out = {}
    for canon, aliases in _COLS.items():
        for cand in aliases:
            if _norm(cand) in norm:
                out[canon] = norm[_norm(cand)]
                break
    return out


def _get(row: dict, cmap: dict, field: str):
    col = cmap.get(field)
    return row.get(col) if col else None


def parse_rate(raw) -> Optional[Decimal]:
    """Tolk en duty expression som ad valorem-brøk. Specifik told → None."""
    if raw is None or raw == "":
        return None
    s = str(raw).strip()
    if re.search(r"(eur|dkk|/|kg|100\s)", s, re.I):  # specifik/sammensat told
        return None
    s = s.replace("%", "").replace(",", ".").strip()
    try:
        val = Decimal(s)
    except (InvalidOperation, ValueError):
        return None
    return val / 100 if val > Decimal("1.5") else val  # 12 → 0.12; 0.12 → 0.12


def _code10(raw) -> str:
    """Normalisér en TARIC GOODS CODE til 10-cifret nøgle (ekstraktet kan have suffiks)."""
    digits = "".join(ch for ch in str(raw or "") if ch.isdigit())
    return digits[:10]


def _aggregate(nom_rows: list[dict], mea_rows: list[dict], groups: dict[str, dict]) -> dict:
    """Kerne-aggregering: nomenklatur + measures (+ grupper) → mfn + arrangements."""
    nmap = _colmap(list(nom_rows[0].keys())) if nom_rows else {}
    has_lang = "lang" in nmap
    has_leaf = "is_leaf" in nmap
    codes: dict[str, dict] = {}
    for r in nom_rows:
        code = _code10(_get(r, nmap, "goods_code"))
        if not code:
            continue
        if has_lang:  # flersproget ekstrakt: vælg foretrukket sprog
            lang = str(_get(r, nmap, "lang") or "").strip().upper()
            existing = codes.get(code, {}).get("_lang")
            if existing in PREFERRED_LANGS and (
                lang not in PREFERRED_LANGS
                or PREFERRED_LANGS.index(lang) >= PREFERRED_LANGS.index(existing)
            ):
                continue
        is_leaf = (str(_get(r, nmap, "is_leaf") or "").strip() in ("1", "1.0", "true", "True")
                   if has_leaf else True)
        codes[code] = {"description": _get(r, nmap, "description") or "", "is_leaf": is_leaf,
                       "_lang": str(_get(r, nmap, "lang") or "").strip().upper() if has_lang else None}

    mmap = _colmap(list(mea_rows[0].keys())) if mea_rows else {}
    mfn: dict[str, Decimal] = {}
    arrangements: dict[str, dict] = {}
    for r in mea_rows:
        code = _code10(_get(r, mmap, "goods_code"))
        mtype = str(_get(r, mmap, "measure_type") or "").strip()
        area = str(_get(r, mmap, "geo_area") or "").strip()
        rate = parse_rate(_get(r, mmap, "duty_rate"))
        if mtype in MFN_MEASURE_TYPES and area.upper() in ERGA_OMNES:
            if rate is not None:
                mfn[code] = rate
        elif mtype in PREFERENCE_MEASURE_TYPES:
            grp = groups.get(area)
            members = grp["members"] if grp else ({area.upper()} if area.isalpha() else set())
            name = grp["name"] if grp else None
            atype = grp["type"] if grp else "fta"
            for country in members:
                cur = arrangements.get(country)
                cand = None if atype == "gsp_standard" else (rate if rate is not None else Decimal(0))
                if cur is None or (cand is not None and cur.get("rate") is not None and cand < cur["rate"]):
                    arrangements[country] = {"name": name, "type": atype}
                    if cand is not None:
                        arrangements[country]["rate"] = float(cand)
    return {"codes": codes, "mfn": mfn, "arrangements": arrangements}


def build_reference(nomenclature: Path, measures: Path, geo_groups: Path) -> dict:
    """Parse de tre fixtur-/CIRCABC-filer til {codes, mfn, arrangements}."""
    grp_rows = _read_table(geo_groups)
    gmap = _colmap(list(grp_rows[0].keys())) if grp_rows else {}
    groups: dict[str, dict] = {}
    for r in grp_rows:
        gid = str(_get(r, gmap, "group_id") or "").strip()
        if not gid:
            continue
        g = groups.setdefault(gid, {
            "name": _get(r, gmap, "arrangement_name"),
            "type": _get(r, gmap, "arrangement_type"), "members": set(),
        })
        member = str(_get(r, gmap, "member_country") or "").strip().upper()
        if member:
            g["members"].add(member)

    nom_rows = _read_table(nomenclature)
    mea_rows = _read_table(measures)
    return _aggregate(nom_rows, mea_rows, groups)


def read_extract(path: Path) -> tuple[list[dict], list[dict]]:
    """Læs et officielt DDS TARIC-ekstrakt (mappe eller .zip) → (nomenklatur, measures)."""
    import glob
    import tempfile
    import zipfile

    if path.suffix.lower() == ".zip":
        tmp = Path(tempfile.mkdtemp())
        with zipfile.ZipFile(path) as z:
            z.extractall(tmp)
        path = tmp
    nom = glob.glob(str(path / "Goods_Nomenclature*.xlsx"))
    mea = glob.glob(str(path / "Measures*.xlsx"))
    if not nom or not mea:
        raise FileNotFoundError(f"Mangler Goods_Nomenclature*/Measures* i {path}")
    return _read_table(Path(nom[0])), _read_table(Path(mea[0]))


def build_from_extract(path: Path) -> dict:
    """Byg referencedata direkte fra et officielt DDS-ekstrakt (mappe/zip).

    Bemærk: DDS-dagsekstrakter er deltaer, og geografiske gruppe-sammensætninger
    ligger i en separat fil i total-ekstraktet — landegruppe-områder (numeriske)
    kan derfor ikke opløses fra et delta alene. Direkte landekoder (GB/IN/...) opløses.
    """
    nom_rows, mea_rows = read_extract(path)
    return _aggregate(nom_rows, mea_rows, groups={})


def build_from_evita(path: Path) -> dict:
    """Byg fuld MFN-tabel fra Skattestyrelsens danske toldtarif-ekstrakt (eVita-XML).

    `toldtarif_ext.xml` (info.skat.dk/download/told/toldtarif_ext.zip) — den fulde
    danske toldtarif: alle deklarérbare varekoder (productLine 80) med officielle
    danske beskrivelser og tredjelands-MFN (measureType 103, ad valorem dutyExpressionId 01).
    Streames med lxml iterparse (hukommelsesbundet, XXE fra). Præferencer pr. land er
    ikke i dette ekstrakt — de kommer fra arrangements (seed eller Trader Export-tot).
    """
    from lxml import etree

    if path.suffix.lower() == ".zip":
        import tempfile, zipfile
        tmp = Path(tempfile.mkdtemp())
        with zipfile.ZipFile(path) as z:
            inner = [n for n in z.namelist() if n.lower().endswith(".xml")][0]
            z.extract(inner, tmp)
            path = tmp / inner

    NS = "{http://www.arcticgroup.se/tariff/arctictariff/export}"
    codes: dict[str, dict] = {}
    mfn: dict[str, Decimal] = {}
    parser = etree.iterparse(str(path), events=("end",), tag=NS + "gono",
                             resolve_entities=False, no_network=True)
    for _, g in parser:
        code = _code10(g.findtext(NS + "goodsNomenclatureItemId"))
        is_leaf = g.findtext(NS + "productLine") == "80"
        desc = None
        for dp in g.iter(NS + "descriptionPeriod"):
            if dp.findtext(NS + "languageId") == "DA":
                desc = dp.findtext(NS + "description")
        if code:
            codes[code] = {"description": (desc or "").strip(), "is_leaf": is_leaf}
            for m in g.iter(NS + "measure"):
                if m.findtext(NS + "measureType") != "103":
                    continue
                for comp in m.iter(NS + "measureComponent"):
                    if comp.findtext(NS + "dutyExpressionId") == "01":  # ad valorem %
                        da = comp.findtext(NS + "dutyAmount")
                        if da not in (None, ""):
                            try:
                                mfn[code] = Decimal(da) / 100  # "12.0" -> 0.12
                            except Exception:
                                pass
        g.clear()
        while g.getprevious() is not None:
            del g.getparent()[0]
    return {"codes": codes, "mfn": mfn, "arrangements": {}}


def download_latest(dest: Path) -> Path:
    """Hent det nyeste officielle TARIC-ekstrakt-zip fra DDS (ingen login).

    NB: Kører via urllib; i nogle sandkasse-miljøer er udgående netværk fra Python
    blokeret — kør da `tools/fetch_taric.sh` (curl) eller angiv --extract manuelt.
    """
    import re
    import urllib.parse
    import urllib.request

    ua = {"User-Agent": "Mozilla/5.0"}
    listing = f"{DDS_BASE}/daily_publications.jsp?Lang=en&Domain=TARIC"
    html = urllib.request.urlopen(urllib.request.Request(listing, headers=ua), timeout=60).read().decode("utf-8", "ignore")
    m = re.search(r"publicationDate=([0-9]{4}-[0-9]{2}-[0-9]{2} [0-9]{2}:[0-9]{2})&message=extract", html)
    if not m:
        raise RuntimeError("Kunne ikke finde nyeste publikation på DDS.")
    pub = m.group(1)
    url = f"{DDS_BASE}/taric_management.jsp?publicationDate={urllib.parse.quote(pub)}&message=extract"
    dest.parent.mkdir(parents=True, exist_ok=True)
    data = urllib.request.urlopen(urllib.request.Request(url, headers=ua), timeout=120).read()
    dest.write_bytes(data)
    return dest


def write_reference(ref: dict, out_dir: Path) -> tuple[int, int]:
    out_dir.mkdir(parents=True, exist_ok=True)
    leaf_rows = [
        {"hs_code": code, "description": meta["description"], "mfn_rate": str(ref["mfn"][code])}
        for code, meta in sorted(ref["codes"].items())
        if meta["is_leaf"] and code in ref["mfn"]
    ]
    with (out_dir / "mfn_rates.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["hs_code", "description", "mfn_rate"])
        w.writeheader()
        w.writerows(leaf_rows)
    # Skriv kun arrangements.json hvis ekstraktet faktisk indeholder præferencer
    # (fx eVita-MFN-ekstraktet gør ikke) — ellers ville en tom fil overskygge seed'et.
    if ref["arrangements"]:
        (out_dir / "arrangements.json").write_text(
            json.dumps({"_note": "Genereret af tools/sync_taric.py fra officielt TARIC-ekstrakt.",
                        "arrangements": ref["arrangements"], "no_preference": []},
                       ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    return len(leaf_rows), len(ref["arrangements"])


def main(argv=None) -> int:
    p = argparse.ArgumentParser(description="Synk TARIC-ekstrakt til tariferingslaget.")
    p.add_argument("--evita", help="dansk toldtarif-ekstrakt (toldtarif_ext.xml eller .zip) — fuld MFN")
    p.add_argument("--extract", help="officielt DDS TARIC-ekstrakt (mappe eller .zip)")
    p.add_argument("--fetch-latest", action="store_true", help="auto-hent nyeste DDS-ekstrakt")
    p.add_argument("--nomenclature")
    p.add_argument("--measures")
    p.add_argument("--geo-groups")
    p.add_argument("--out", default=str(REFERENCE))
    p.add_argument("--check", action="store_true", help="verificér at referencedata findes")
    args = p.parse_args(argv)

    if args.check:
        ok = (REFERENCE / "mfn_rates.csv").exists() or (REFERENCE / "seed_mfn_rates.csv").exists()
        print("OK — tarif-referencedata til stede." if ok else "MANGLER — ingen tarif-referencedata.")
        return 0 if ok else 1

    if args.evita:
        ref = build_from_evita(Path(args.evita))
    elif args.fetch_latest:
        zip_path = download_latest(Path(args.out) / "taric_raw" / "latest.zip")
        print(f"Hentet: {zip_path}")
        ref = build_from_extract(zip_path)
    elif args.extract:
        ref = build_from_extract(Path(args.extract))
    elif args.nomenclature and args.measures and args.geo_groups:
        ref = build_reference(Path(args.nomenclature), Path(args.measures), Path(args.geo_groups))
    else:
        print("Vælg én: --fetch-latest | --extract <mappe/zip> | "
              "--nomenclature/--measures/--geo-groups | --check.\n"
              "Officielt TARIC-ekstrakt: DG TAXUD DDS (gratis, ingen login) — se tools/fetch_taric.sh.")
        return 2

    codes, arrs = write_reference(ref, Path(args.out))
    print(f"Synket: {codes} deklarérbare koder m. MFN-sats, {arrs} lande m. præference.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
