"""Indlæser for generiske Excel/CSV-udtræk af deklarationslinjer.

Realistisk det format, de første kundedata lander i (ERP-/told-udtræk). Adapteren
oversætter en flad tabel til de samme analyseklare rækker som
``Declaration.to_rows()``, så analyselaget har én indgang uanset kilde.

Kolonnenavne matches fleksibelt (dansk/engelsk, store/små bogstaver, mellemrum vs.
underscore). Beløb coerces til ``Decimal``; effektiv toldsats udledes hvis den mangler.
Filen modificeres aldrig — den læses og kasseres af kalderen.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional, Union

# Kanonisk feltnavn -> accepterede kolonne-aliaser (normaliseret: lower, uden
# mellemrum/underscore/bindestreg).
_ALIASES = {
    "issue_datetime": ["issuedatetime", "date", "dato", "antagelsesdato", "afsendelsesdato"],
    "consignor_name": ["consignorname", "consignor", "leverandør", "leverandor", "supplier", "afsender"],
    "importer_eori": ["importereori", "importør", "importer", "modtager"],
    "origin_country": ["origincountry", "oprindelsesland", "origin", "countryoforigin"],
    "commodity_code": ["commoditycode", "varekode", "hscode", "hs", "tariffclassification", "varenummer"],
    "description": ["description", "varebeskrivelse", "beskrivelse", "vare"],
    "customs_value_dkk": ["customsvaluedkk", "customsvalue", "toldværdi", "toldvaerdi", "statistiskværdi", "statistiskvaerdi", "statisticalvalue"],
    "customs_duty": ["customsduty", "importduty", "told", "duty", "importtold"],
    "item_invoice_amount": ["iteminvoiceamount", "varenspris", "fakturaværdi", "fakturavaerdi", "invoiceamount"],
    "invoice_currency": ["invoicecurrency", "valuta", "currency", "møntsort", "moentsort"],
    "net_mass": ["netmass", "nettovægt", "nettovaegt", "nettomasse", "netweight", "netweightkgm"],
    "gross_mass": ["grossmass", "bruttovægt", "bruttovaegt", "bruttomasse", "grossweight"],
    "border_mot": ["bordermot", "motborder", "transportmådegrænse", "transportmaadegraense", "modeoftransport", "mot"],
    "inland_mot": ["inlandmot", "motinland", "transportmådeindland", "transportmaadeindland", "indenlandsktransportmåde", "indenlandsktransportmaade"],
    "cpc": ["cpc", "procedurekode", "customsprocedurecode"],
    "duty_regime_code": ["dutyregimecode", "præference", "praeference", "preference", "preferencerate"],
    "shipments": ["shipments", "forsendelser", "antalforsendelser"],
}

_NUMERIC = {
    "customs_value_dkk", "customs_duty", "item_invoice_amount",
    "net_mass", "gross_mass", "shipments",
}


def _norm(name: str) -> str:
    return "".join(ch for ch in name.lower() if ch.isalnum())


def _build_columnmap(header: list[str]) -> dict[str, str]:
    """Kortlæg kanonisk feltnavn -> faktisk kolonnenavn i tabellen."""
    norm_to_actual = {_norm(h): h for h in header}
    mapping: dict[str, str] = {}
    for canon, aliases in _ALIASES.items():
        for cand in [canon] + aliases:
            if _norm(cand) in norm_to_actual:
                mapping[canon] = norm_to_actual[_norm(cand)]
                break
    return mapping


def _to_decimal(raw: str) -> Optional[Decimal]:
    if raw is None:
        return None
    s = str(raw).strip().replace(" ", "").replace(" ", "")
    if not s:
        return None
    # Håndtér dansk talformat (1.234.567,89) og kr.-suffiks.
    s = s.replace("kr.", "").replace("kr", "").replace("%", "").strip()
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _coerce_row(raw: dict, colmap: dict[str, str]) -> dict:
    row: dict = {}
    for canon, actual in colmap.items():
        val = raw.get(actual)
        if canon in _NUMERIC:
            row[canon] = _to_decimal(val)
        else:
            row[canon] = (str(val).strip() if val not in (None, "") else None)
    # Udled effektiv toldsats hvis told + værdi findes.
    value = row.get("customs_value_dkk")
    duty = row.get("customs_duty")
    if value and duty is not None:
        row["effective_duty_rate"] = duty / value
    return row


def parse_tabular(source: Union[str, bytes, Path], filename: str = "") -> list[dict]:
    """Læs CSV eller XLSX til analyseklare rækker. Format afgøres af filendelse/indhold."""
    name = (filename or (str(source) if isinstance(source, (str, Path)) else "")).lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(source)
    return _parse_csv(source)


def _read_bytes(source: Union[str, bytes, Path]) -> bytes:
    if isinstance(source, bytes):
        return source
    return Path(str(source)).read_bytes()


def _parse_csv(source: Union[str, bytes, Path]) -> list[dict]:
    if isinstance(source, (str, Path)) and Path(str(source)).exists():
        text = Path(str(source)).read_text(encoding="utf-8-sig")
    elif isinstance(source, bytes):
        text = source.decode("utf-8-sig")
    else:
        text = str(source)
    # Gæt afgrænser (komma eller semikolon).
    sample = text[:2048]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delim)
    colmap = _build_columnmap(reader.fieldnames or [])
    return [_coerce_row(r, colmap) for r in reader]


def _parse_xlsx(source: Union[str, bytes, Path]) -> list[dict]:
    from openpyxl import load_workbook  # lazy: kun krævet for xlsx

    data = _read_bytes(source)
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    colmap = _build_columnmap(header)
    out = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue
        raw = {header[i]: values[i] for i in range(min(len(header), len(values)))}
        out.append(_coerce_row(raw, colmap))
    wb.close()
    return out
